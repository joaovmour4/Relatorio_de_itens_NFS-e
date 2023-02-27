from tkinter import *
from tkinter.filedialog import *
import xml.etree.ElementTree as Et
import openpyxl
from openpyxl.styles.borders import Border


def collect_arquivo(entry):
    entry.insert(0, askopenfilename())


def save_arquivo(entry):
    entry.insert(0, askdirectory())


def executar(arquivo, directory):
    wb = openpyxl.Workbook()
    sheet = wb.active

    xml = Et.parse(f'{arquivo}')

    root = xml.getroot()
    i = 1
    blueFill = openpyxl.styles.PatternFill(start_color='99CCFF',
                                           end_color='99CCFF',
                                           fill_type='solid')

    blue1Fill = openpyxl.styles.PatternFill(start_color='CCECFF',
                                            end_color='CCECFF',
                                            fill_type='solid')

    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 16
    sheet.column_dimensions['F'].width = 25

    for nota in root:
        cll_nota = sheet.cell(row=i, column=2)
        sheet.merge_cells(f'B{i}:F{i}')
        cll_nota.alignment = openpyxl.styles.Alignment(horizontal='center')
        cll_nota.value = f'nota {nota.find("tcInfNFE")[0].text}'
        cll_nota.fill = blueFill
        i += 1
        for elemento in nota:
            if elemento.tag == 'tcInfNFE':
                cll_elemento = sheet.cell(row=i, column=2)
                for campo in elemento:
                    if campo.tag == 'tcInfItens':
                        cll_campo = sheet.cell(row=i, column=3)
                        cll_campo.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                        for item in campo:
                            j = 2
                            cll_item = sheet.cell(row=i, column=j)
                            cll_item.fill = blue1Fill
                            sheet.merge_cells(f'B{i}:F{i}')
                            cll_item.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                            cll_item.value = f'Informações do Item'
                            i += 1
                            for detalhe in item:
                                cll_detalhe_info = sheet.cell(row=i, column=j)
                                cll_detalhe = sheet.cell(row=i + 1, column=j)
                                cll_detalhe.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                                  vertical='center', wrapText=True)
                                cll_detalhe_info.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                                       vertical='center')
                                j += 1
                                match str(detalhe.tag):
                                    case 'tsSeqItem':
                                        cll_detalhe_info.value = f'ID item'
                                    case 'tsQtdItem':
                                        cll_detalhe_info.value = f'Quantidade'
                                    case 'tsDesItem':
                                        cll_detalhe_info.value = f'Descrição'
                                    case 'tsVlrItem':
                                        cll_detalhe_info.value = f'Valor do item'
                                    case 'tsTotItem':
                                        cll_detalhe_info.value = f'Valor total dos itens'
                                cll_detalhe.value = f'{detalhe.text}'
                            i += 3
                        i += 2
        i += 2

    wb.save(f'{directory}/Relatorio-NFS-e_{root[0][0][3].text}.xlsx')


janela = Tk()
janela.geometry('600x200')
janela.title('Gerar relatório de NFS-e')

Label(janela, text='Diretório do arquivo XML:').grid(column=0, row=0, sticky='w', padx=50, ipady=10)
entry_arquivo = Entry(janela)
entry_arquivo.grid(column=1, row=0, padx=10)

Label(janela, text='Diretório de salvamento do relatório:').grid(column=0, row=1, padx=50)
entry_save = Entry(janela)
entry_save.grid(column=1, row=1, padx=10)

Button(janela, text='...', command=lambda: collect_arquivo(entry_arquivo)).grid(column=2, row=0)
Button(janela, text='...', command=lambda: save_arquivo(entry_save)).grid(column=2, row=1)
Button(janela, text='Executar', command=lambda: executar(entry_arquivo.get(), entry_save.get())).grid(column=3, row=3)

janela.mainloop()
